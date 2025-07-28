import json
import openpyxl


def anonymize_dispersal_spreadsheet(file_name, output_file_name):
    wb = openpyxl.load_workbook(file_name)

    actors_sheet = wb["model-v1-actors"]
    events_sheet = wb["model-v1-events"]

    actor_id_column = "A"
    actor_name_column = "B"
    actor_type_column = "C"
    actor_address_1_column = "F"
    actor_postcode_column = "J"

    recipient_id_column = "AB"
    recipient_name_column = "AC"

    actors = {}

    # Anonymize actors sheet
    for row in range(2, actors_sheet.max_row + 1):
        actor_type = actors_sheet[f"{actor_type_column}{row}"].value
        if actor_type == "individual":
            actor_id = actors_sheet[f"{actor_id_column}{row}"].value
            actor_name = actors_sheet[f"{actor_name_column}{row}"].value

            index = len(actors)
            new_name = f"person{index}"
            new_id = f"p{index}"

            actors[actor_id] = {"id": new_id, "name": new_name}

            actors_sheet[f"{actor_id_column}{row}"].value = new_id
            actors_sheet[f"{actor_name_column}{row}"].value = new_name
            actors_sheet[f"{actor_address_1_column}{row}"].value = ""
            actors_sheet[f"{actor_postcode_column}{row}"].value = ""

    # Update actor ids in "events" sheet
    for row in range(2, events_sheet.max_row + 1):
        recipient_id = events_sheet[f"{recipient_id_column}{row}"].value
        if recipient_id in actors:
            events_sheet[f"{recipient_id_column}{row}"].value = actors[recipient_id][
                "id"
            ]
            events_sheet[f"{recipient_name_column}{row}"].value = actors[recipient_id][
                "name"
            ]

    # Save the changes to a new file
    wb.save(output_file_name)


if __name__ == "__main__":
    with open("config.json") as f:
        config = json.load(f)
        dispersal_sheet_file = config["dispersal_sheet_file"]
        dispersal_sheet_anon = config["dispersal_sheet_anon"]

anonymize_dispersal_spreadsheet(dispersal_sheet_file, dispersal_sheet_anon)
